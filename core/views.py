from django.shortcuts import render, redirect
from django.db import transaction
from django.http import HttpResponse
from .forms import TarifeYuklemeForm, FiyatListesiOlusturmaForm
from .services import tarifeleri_karsilastir, fiyat_listesi_hazirla
from .models import HizmetListesi, TarifeKarsilastirma
import pandas as pd
from decimal import Decimal
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# GÜNCELLENEN GRUP TANIMLARI
GRUP_TANIMLARI = {
    'İHR': 'İHRACAT İŞLEMLERİ TÜM GÜMRÜKLER İÇİN (HER BEYANNAME)',
    'İTH': 'İTHALAT İŞLEMLERİ (HER BEYANNAME)',
    'TR': 'TRANSİT İŞLEMLERİ',
    'ANT': 'GÜMRÜK ANTREPO İŞLEMLERİ',
    'DAN': 'DANIŞMANLIK ÜCRETLERİ',
    'ÖZ': 'ÖZELLİK ARZ EDEN İŞLEMLER',
    'TRM': 'TARIM İŞLEMLERİ',
    'TSE': 'TSE/TAREKS İŞLEMLERİ',
    'Uİ': 'UZLAŞMA & İTİRAZ İŞLEMLERİ',
    'ODİ': 'OKUYAN DİĞER İŞLEMLER'
}

def fiyat_duzelt(deger):
    """Excel'den gelen veriyi sayıya çevirir."""
    if pd.isna(deger) or str(deger).strip() in ['-', '']:
        return 0.0
    
    if isinstance(deger, (int, float)):
        return float(deger)

    try:
        s = str(deger).replace('TL', '').replace('tl', '').replace('₺', '').strip()
        s = s.replace('.', '').replace(',', '.')
        return float(s)
    except:
        return 0.0

def anasayfa(request):
    """Fiyat Oluşturma ve Excel İndirme Sayfası"""
    veriler = []
    hata = None
    form = FiyatListesiOlusturmaForm(request.POST or None, request.FILES or None)

    if request.method == 'POST':
        # --- DURUM 1: EXCEL İNDİR BUTONUNA BASILDIYSA ---
        if 'excel_indir' in request.POST:
            try:
                # 1. HTML Formundan Verileri Çek
                kodlar = request.POST.getlist('liste_kod[]')
                hizmetler = request.POST.getlist('liste_hizmet[]')
                fiyatlar = request.POST.getlist('liste_fiyat[]')
                durumlar = request.POST.getlist('liste_durum[]')
                eski_fiyatlar = request.POST.getlist('liste_eski_fiyat[]')
                yil = request.POST.get('hedef_yil_hidden', '2025')

                # 2. Verileri Eşleşen ve Eşleşmeyen Diye Ayır
                eslesen_rows = []
                eslesmeyen_rows = []

                for k, h, f, d, ef in zip(kodlar, hizmetler, fiyatlar, durumlar, eski_fiyatlar):
                    if not k or k == '-' or k == 'nan':
                        continue

                    item = {
                        'kod': str(k).strip(),
                        'hizmet': h,
                        'fiyat': f,
                        'eski_fiyat': ef
                    }

                    if d == 'Eşleşemeyen':
                        eslesmeyen_rows.append(item)
                    else:
                        eslesen_rows.append(item)

                # 3. Excel Workbook Oluştur
                wb = Workbook()
                ws = wb.active
                ws.title = "Fiyat Listesi"

                # -- STİLLER --
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                group_font = Font(bold=True)
                group_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                unmatched_header_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                # Sütun Genişlikleri
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 75
                ws.column_dimensions['C'].width = 25

                # --- BÖLÜM 1: EŞLEŞENLER (NORMAL LİSTE) ---
                headers = ['KOD', 'HİZMET KONUSU', f'{yil} YILI ÜCRETLENDİRME']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border

                son_grup = None

                for row_data in eslesen_rows:
                    kod = row_data['kod']
                    prefix = kod.split('-')[0] if '-' in kod else kod
                    grup_kodu = 'ÖZ' if prefix in ['ÖZ', 'SB'] else prefix

                    if grup_kodu != son_grup:
                        aciklama = GRUP_TANIMLARI.get(grup_kodu, f'{grup_kodu} İşlemleri')
                        ws.append([grup_kodu, aciklama, ''])
                        current_row = ws.max_row
                        for col in range(1, 4):
                            cell = ws.cell(row=current_row, column=col)
                            cell.fill = group_fill
                            cell.font = group_font
                            cell.border = thin_border
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
                        son_grup = grup_kodu

                    # SIFIR KONTROLÜ: Fiyat 0 ise boş bırak
                    fiyat_val = row_data['fiyat']
                    try:
                        if float(str(fiyat_val).replace(',', '')) == 0:
                            fiyat_val = ""
                    except:
                        pass

                    # Veri Satırı Ekle
                    ws.append([row_data['kod'], row_data['hizmet'], fiyat_val])
                    current_row = ws.max_row
                    for col in range(1, 4):
                         cell = ws.cell(row=current_row, column=col)
                         cell.border = thin_border
                         # METNİ KAYDIR (Hizmet konusu için 2. sütuna uygula)
                         if col == 2:
                             cell.alignment = Alignment(wrap_text=True, vertical='center')
                         elif col == 3: # Fiyatları sağa yasla
                             cell.alignment = Alignment(horizontal='right', vertical='center')
                         else:
                             cell.alignment = Alignment(vertical='center')


                # --- BÖLÜM 2: EŞLEŞEMEYENLER (ALT KISIM) ---
                if eslesmeyen_rows:
                    ws.append([]) 
                    ws.append([]) 
                    unmatched_headers = ['KOD', 'HİZMET KONUSU', 'ESKİ ÜCRETLENDİRME']
                    ws.append(unmatched_headers)

                    current_row = ws.max_row
                    for col in range(1, 4):
                        cell = ws.cell(row=current_row, column=col)
                        cell.fill = unmatched_header_fill
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    for row_data in eslesmeyen_rows:
                        # SIFIR KONTROLÜ: Eski Fiyat 0 ise boş bırak
                        eski_fiyat_val = row_data['eski_fiyat']
                        try:
                            if float(str(eski_fiyat_val).replace(',', '')) == 0:
                                eski_fiyat_val = ""
                        except:
                            pass

                        ws.append([row_data['kod'], row_data['hizmet'], eski_fiyat_val])
                        current_row = ws.max_row
                        for col in range(1, 4):
                             cell = ws.cell(row=current_row, column=col)
                             cell.border = thin_border
                             if col == 2:
                                 cell.alignment = Alignment(wrap_text=True, vertical='center')
                             elif col == 3:
                                 cell.alignment = Alignment(horizontal='right', vertical='center')
                             else:
                                 cell.alignment = Alignment(vertical='center')

                # 5. Çıktıyı Hazırla
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename=Fiyat_Listesi_{yil}.xlsx'
                return response

            except Exception as e:
                hata = f"Excel oluşturulurken hata: {str(e)}"

        # --- DURUM 2: HESAPLA BUTONU ---
        elif form.is_valid():
            hedef_yil = form.cleaned_data['yil']
            dosya = request.FILES.get('dosya')
            try:
                path = dosya if dosya else None
                ham_liste = fiyat_listesi_hazirla(hedef_yil, path)
                gosterim_listesi = []
                son_grup = None

                for satir in ham_liste:
                    kod = str(satir['kod']).strip()
                    if not kod or kod == '-': continue

                    prefix = kod.split('-')[0] if '-' in kod else kod
                    grup_kodu = 'ÖZ' if prefix in ['ÖZ', 'SB'] else prefix

                    if grup_kodu != son_grup:
                        aciklama = GRUP_TANIMLARI.get(grup_kodu, f'{grup_kodu} İşlemleri')
                        gosterim_listesi.append({
                            'is_header': True,
                            'kod': grup_kodu,
                            'aciklama': aciklama
                        })
                        son_grup = grup_kodu
                    satir['is_header'] = False
                    gosterim_listesi.append(satir)

                veriler = gosterim_listesi
                if not veriler:
                    hata = f"{hedef_yil} verisi bulunamadı."

            except Exception as e:
                hata = f"Hesaplama hatası: {str(e)}"

    return render(request, 'core/fiyat_olusturma.html', {
        'form': form,
        'veriler': veriler,
        'hata': hata
    })

def tarife_karsilastirma(request):
    veriler = None
    hata = None
    form = TarifeYuklemeForm()

    if request.method == 'POST':
        if 'analiz_et' in request.POST:
            form = TarifeYuklemeForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    tarife_obj = form.save()
                    ham_veriler, hata = tarifeleri_karsilastir(tarife_obj.eski_yil_dosyasi.path, tarife_obj.yeni_yil_dosyasi.path)
                    if ham_veriler:
                        request.session['gecici_veriler'] = ham_veriler
                        request.session['analiz_yili'] = 2025

                        gosterim_listesi = []
                        son_grup = None

                        for satir in ham_veriler:
                            yeni_kod = str(satir['Yeni_Kod']).strip()
                            prefix = yeni_kod.split('-')[0] if '-' in yeni_kod else yeni_kod
                            grup_kodu = 'ÖZ' if prefix in ['ÖZ', 'SB'] else prefix

                            if grup_kodu != son_grup:
                                aciklama = GRUP_TANIMLARI.get(grup_kodu, f'{grup_kodu} İşlemleri')
                                gosterim_listesi.append({
                                    'is_header': True,
                                    'kod': grup_kodu,
                                    'aciklama': aciklama
                                })
                                son_grup = grup_kodu
                            satir['is_header'] = False
                            gosterim_listesi.append(satir)
                        veriler = gosterim_listesi

                except Exception as e:
                    hata = f"Analiz hatası: {str(e)}"

        elif 'kaydet' in request.POST:
            gecici_veriler = request.session.get('gecici_veriler')
            analiz_yili = request.session.get('analiz_yili', 2025)

            if gecici_veriler:
                try:
                    with transaction.atomic():
                        for satir in gecici_veriler:
                            yeni_obj, _ = HizmetListesi.objects.get_or_create(
                                yil=analiz_yili,
                                hizmet_kodu=satir['Yeni_Kod'],
                                defaults={
                                    'hizmet_adi': satir['Yeni_Hizmet'],
                                    'tutar': satir['Yeni_Fiyat'].replace(',', '') if satir['Yeni_Fiyat'] != '-' else 0
                                }
                            )

                            eski_obj = None
                            if satir['Eski_Kod'] != '-':
                                eski_obj, _ = HizmetListesi.objects.get_or_create(
                                    yil=analiz_yili - 1,
                                    hizmet_kodu=satir['Eski_Kod'],
                                    defaults={
                                        'hizmet_adi': satir['Eski_Hizmet'],
                                        'tutar': satir['Eski_Fiyat'].replace(',', '') if satir['Eski_Fiyat'] != '-' else 0
                                    }
                                )

                            kayit_var_mi = TarifeKarsilastirma.objects.filter(yeni_hizmet=yeni_obj).exists()

                            if not kayit_var_mi:
                                TarifeKarsilastirma.objects.create(
                                    yeni_hizmet=yeni_obj,
                                    eski_hizmet=eski_obj,
                                    tutar_fark=float(satir['Fark'].replace(',', '')) if satir['Fark'] != '-' else 0,
                                    yuzde_degisim=float(satir['Degisim_Yuzde']) if satir['Degisim_Yuzde'] != '-' else 0,
                                    durum=satir['Durum']
                                )
                    del request.session['gecici_veriler']
                    return render(request, 'core/tarife_karsilastirma.html', {'mesaj': 'Veriler başarıyla veritabanına kaydedildi!', 'form': form})
                except Exception as e:
                    hata = f"Veritabanına kaydederken hata oluştu: {str(e)}"

    return render(request, 'core/tarife_karsilastirma.html', {'form': form, 'veriler': veriler, 'hata': hata})